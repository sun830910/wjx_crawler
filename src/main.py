# -*- coding: utf-8 -*-

"""
Created on 2020-10-30 16:18
@Author  : Justin Jiang
@Email   : jw_jiang@pku.edu.com
"""

from selenium import webdriver
from bs4 import BeautifulSoup
import re
import xlrd
import xlwt
import openpyxl
import time  # 引入time模块


def read_piao(str):
    patt = '>\d+'
    regex = re.search(patt, str)[0].strip('>')
    return regex


def read_name(str):
    patt = '>.*</div'
    regex = re.search(patt, str)[0]
    start = regex.find(">")
    end = regex.find("<")
    return regex[start + 1:end]


def get_data(url):
    chrome_driver_path = '../../chromedriver'
    browser = webdriver.Chrome(executable_path=chrome_driver_path)
    browser.get(url)
    page_source = browser.page_source
    browser.close()
    soup = BeautifulSoup(page_source, 'lxml')
    get_piao = soup.find_all('div', class_='tppiao')
    get_name = soup.find_all('div', class_='tpchoice')

    piao_result = []
    user_result = []
    print("共有 {} 个参赛者".format(len(get_piao)))
    for idx, num in enumerate(get_piao):
        piao_num = read_piao(str(num))
        user_name = read_name(str(get_name[idx]))
        print("第 {} 名  || {} || 票数： {} 票".format(idx + 1, user_name, piao_num))
        piao_result.append(piao_num)
        user_result.append(user_name)
    return piao_result, user_result


def write_excel_xlsx(path, users, scores):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for i in range(len(users)):
        sheet.cell(row=1, column=i + 1, value=str(users[i]))
        sheet.cell(row=2, column=i + 1, value=str(scores[i]))
    workbook.save(path)
    print("xlsx格式表格写入数据成功！")


if __name__ == '__main__':
    url = 'https://tp.wjx.top/wjx/join/tpresult.aspx?activity=95463666'
    scores, users = get_data(url)
    file_name = '../data/' + time.strftime("%H_%M_%S", time.localtime()) + '.xlsx'
    print(file_name)
    print('当前时间:' + time.strftime("%H:%M:%S", time.localtime()))
    write_excel_xlsx(file_name, users, scores)
